"""
Dynamic Excel import/export functionality for custom catalogue structures
"""

import pandas as pd
from typing import List, Dict, Any
import os
from src.gui.setup_wizard import ColumnDefinition

class DynamicExcelHandler:
    """Handles Excel import and export operations for dynamic catalogues"""
    
    def __init__(self, catalogue_config: Dict):
        """Initialize Excel handler with catalogue configuration
        
        Args:
            catalogue_config: Dictionary containing catalogue name and column definitions
        """
        self.catalogue_config = catalogue_config
        self.columns = [ColumnDefinition.from_dict(col) for col in catalogue_config['columns']]
    
    def import_from_excel(self, file_path: str) -> List[Dict]:
        """Import items from Excel file based on dynamic column structure
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of item dictionaries
            
        Raises:
            Exception: If file cannot be read or format is invalid
        """
        if not os.path.exists(file_path):
            raise Exception(f"File not found: {file_path}")
        
        try:
            # Read Excel file
            df = pd.read_excel(file_path)
            
            # Normalize column names (case-insensitive, strip whitespace)
            df.columns = df.columns.str.strip()
            
            # Create mapping from Excel columns to our column names
            column_mapping = {}
            unmapped_columns = []
            
            for column in self.columns:
                mapped = False
                
                # Try exact match first (case-insensitive)
                for excel_col in df.columns:
                    if excel_col.lower() == column.display_name.lower():
                        column_mapping[excel_col] = column.name
                        mapped = True
                        break
                
                # Try partial matches if exact match fails
                if not mapped:
                    for excel_col in df.columns:
                        if (column.display_name.lower() in excel_col.lower() or 
                            excel_col.lower() in column.display_name.lower()):
                            column_mapping[excel_col] = column.name
                            mapped = True
                            break
                
                if not mapped:
                    unmapped_columns.append(column.display_name)
            
            # Show mapping information
            if unmapped_columns:
                print(f"Warning: Could not map columns: {', '.join(unmapped_columns)}")
            
            # Rename columns based on mapping
            df = df.rename(columns=column_mapping)
            
            # Process each row
            items = []
            for _, row in df.iterrows():
                item = {}
                
                for column in self.columns:
                    if column.name in df.columns:
                        raw_value = row[column.name]
                        
                        # Handle NaN values
                        if pd.isna(raw_value):
                            value = column.default_value or self._get_default_for_type(column.data_type)
                        else:
                            value = self._convert_excel_value(raw_value, column)
                        
                        item[column.name] = value
                    else:
                        # Use default value if column not found in Excel
                        item[column.name] = column.default_value or self._get_default_for_type(column.data_type)
                
                # Only add items that have at least one non-empty required field
                if self._is_valid_item(item):
                    items.append(item)
            
            return items
            
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    def export_to_excel(self, items: List[Dict], file_path: str):
        """Export items to Excel file based on dynamic column structure
        
        Args:
            items: List of item dictionaries
            file_path: Path where to save the Excel file
            
        Raises:
            Exception: If file cannot be written
        """
        try:
            if not items:
                # Create empty DataFrame with column headers
                data = {col.display_name: [] for col in self.columns}
                df = pd.DataFrame(data)
            else:
                # Convert items to DataFrame
                data = {}
                for column in self.columns:
                    column_data = []
                    for item in items:
                        raw_value = item.get(column.name, column.default_value)
                        formatted_value = self._format_value_for_excel(raw_value, column)
                        column_data.append(formatted_value)
                    
                    data[column.display_name] = column_data
                
                df = pd.DataFrame(data)
            
            # Create Excel writer with formatting
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Catalogue', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Catalogue']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, max 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format header row
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Add data validation for boolean columns
                from openpyxl.worksheet.datavalidation import DataValidation
                
                for col_idx, column in enumerate(self.columns, 1):
                    if column.data_type == 'boolean':
                        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
                        dv.error = 'Please select Yes or No'
                        dv.errorTitle = 'Invalid Entry'
                        
                        # Apply to the entire column (up to row 1000)
                        col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                        dv.add(f'{col_letter}2:{col_letter}1000')
                        worksheet.add_data_validation(dv)
            
        except Exception as e:
            raise Exception(f"Error writing Excel file: {str(e)}")
    
    def validate_excel_format(self, file_path: str) -> Dict[str, Any]:
        """Validate Excel file format and return information about it
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dictionary with validation results
        """
        try:
            df = pd.read_excel(file_path)
            
            result = {
                'valid': True,
                'rows': len(df),
                'columns': list(df.columns),
                'matched_columns': [],
                'unmatched_columns': [],
                'missing_columns': [],
                'errors': []
            }
            
            # Normalize column names
            df.columns = df.columns.str.strip()
            
            # Check which columns can be mapped
            for column in self.columns:
                matched = False
                for excel_col in df.columns:
                    if (excel_col.lower() == column.display_name.lower() or
                        column.display_name.lower() in excel_col.lower() or
                        excel_col.lower() in column.display_name.lower()):
                        result['matched_columns'].append({
                            'catalogue_column': column.display_name,
                            'excel_column': excel_col,
                            'required': column.required
                        })
                        matched = True
                        break
                
                if not matched:
                    if column.required:
                        result['missing_columns'].append(column.display_name)
                    else:
                        result['unmatched_columns'].append(column.display_name)
            
            # Check for required columns
            if result['missing_columns']:
                result['valid'] = False
                result['errors'].append(f"Missing required columns: {', '.join(result['missing_columns'])}")
            
            return result
            
        except Exception as e:
            return {
                'valid': False,
                'rows': 0,
                'columns': [],
                'matched_columns': [],
                'unmatched_columns': [],
                'missing_columns': [],
                'errors': [f"Error reading file: {str(e)}"]
            }
    
    def _convert_excel_value(self, value: Any, column: ColumnDefinition) -> Any:
        """Convert Excel value to appropriate type for the column
        
        Args:
            value: Raw value from Excel
            column: Column definition
            
        Returns:
            Converted value
        """
        if pd.isna(value) or value == '':
            return column.default_value or self._get_default_for_type(column.data_type)
        
        try:
            if column.data_type == 'number':
                return int(float(value))  # Convert via float to handle "1.0" -> 1
            elif column.data_type == 'decimal':
                return float(value)
            elif column.data_type == 'boolean':
                if isinstance(value, bool):
                    return value
                elif isinstance(value, str):
                    return value.lower() in ('true', '1', 'yes', 'on', 'y')
                else:
                    return bool(value)
            elif column.data_type == 'date':
                if pd.isna(value):
                    return column.default_value or ''
                # Handle various date formats
                if isinstance(value, pd.Timestamp):
                    return value.strftime('%Y-%m-%d')
                else:
                    return str(value)
            else:  # text
                return str(value).strip()
        except (ValueError, TypeError):
            # If conversion fails, use default value
            return column.default_value or self._get_default_for_type(column.data_type)
    
    def _format_value_for_excel(self, value: Any, column: ColumnDefinition) -> Any:
        """Format value for Excel export
        
        Args:
            value: Raw value
            column: Column definition
            
        Returns:
            Formatted value
        """
        if value is None or value == '':
            return ''
        
        if column.data_type == 'boolean':
            if isinstance(value, bool):
                return 'Yes' if value else 'No'
            elif isinstance(value, (int, str)):
                return 'Yes' if bool(value) else 'No'
            else:
                return 'No'
        elif column.data_type == 'decimal':
            try:
                return float(value)
            except (ValueError, TypeError):
                return 0.0
        elif column.data_type == 'number':
            try:
                return int(value)
            except (ValueError, TypeError):
                return 0
        else:
            return str(value)
    
    def _get_default_for_type(self, data_type: str) -> Any:
        """Get default value for a data type
        
        Args:
            data_type: Column data type
            
        Returns:
            Default value
        """
        defaults = {
            'text': '',
            'number': 0,
            'decimal': 0.0,
            'date': '',
            'boolean': False
        }
        return defaults.get(data_type, '')
    
    def _is_valid_item(self, item: Dict) -> bool:
        """Check if an item has valid data (at least one required field filled)
        
        Args:
            item: Item dictionary
            
        Returns:
            True if item is valid
        """
        # Check if any required field has a non-empty value
        required_columns = [col for col in self.columns if col.required]
        
        if not required_columns:
            # If no required columns, check if any field has a non-empty value
            return any(str(value).strip() for value in item.values() if value is not None)
        
        # Check required columns
        for column in required_columns:
            value = item.get(column.name)
            if value is not None and str(value).strip():
                return True
        
        return False
    
    def create_template_excel(self, file_path: str):
        """Create an Excel template file with column headers and sample data
        
        Args:
            file_path: Path where to save the template file
        """
        try:
            # Create sample data
            sample_data = {}
            for column in self.columns:
                sample_data[column.display_name] = [self._get_sample_value(column)]
            
            df = pd.DataFrame(sample_data)
            
            # Create Excel file with formatting
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Template', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Template']
                
                # Format header row
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max(max_length + 2, 15), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add instructions in a separate sheet
                instructions_data = {
                    'Column Name': [col.display_name for col in self.columns],
                    'Data Type': [col.data_type.title() for col in self.columns],
                    'Required': ['Yes' if col.required else 'No' for col in self.columns],
                    'Default Value': [col.default_value or '' for col in self.columns],
                    'Description': [self._get_column_description(col) for col in self.columns]
                }
                
                instructions_df = pd.DataFrame(instructions_data)
                instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
                
                # Format instructions sheet
                inst_worksheet = writer.sheets['Instructions']
                for cell in inst_worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                for column in inst_worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max(max_length + 2, 15), 60)
                    inst_worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            raise Exception(f"Error creating template file: {str(e)}")
    
    def _get_sample_value(self, column: ColumnDefinition) -> str:
        """Get a sample value for a column
        
        Args:
            column: Column definition
            
        Returns:
            Sample value as string
        """
        if column.default_value:
            return column.default_value
        
        samples = {
            'text': f'Sample {column.display_name}',
            'number': '10',
            'decimal': '10.50',
            'date': '2024-01-01',
            'boolean': 'Yes'
        }
        return samples.get(column.data_type, 'Sample Value')
    
    def _get_column_description(self, column: ColumnDefinition) -> str:
        """Get description for a column
        
        Args:
            column: Column definition
            
        Returns:
            Column description
        """
        descriptions = {
            'text': 'Enter text/string values',
            'number': 'Enter whole numbers (integers)',
            'decimal': 'Enter decimal numbers',
            'date': 'Enter dates in YYYY-MM-DD format',
            'boolean': 'Enter Yes/No or True/False'
        }
        
        base_desc = descriptions.get(column.data_type, 'Enter appropriate value')
        if column.required:
            base_desc += ' (Required)'
        
        return base_desc
